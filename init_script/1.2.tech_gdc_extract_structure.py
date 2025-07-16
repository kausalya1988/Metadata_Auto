import yaml
from ruamel.yaml import YAML
import openpyxl
import os
import pandas as pd
import warnings
import re

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

def load_list_of_value(workbook):
    """Load the columns accepted values"""
    lov = workbook["List of value"]
    
    if lov:
        lov_df = pd.DataFrame(lov.values)
        if lov_df.shape[1] >= 3:
            lov_df = lov_df.iloc[:, :3]
            lov_df.columns = ['FIELD', 'LOV', 'COUNTRY']
            lov_df = lov_df.iloc[1:]        
            lov_grouped_df = lov_df.groupby(['FIELD', 'COUNTRY'])['LOV'].agg(list).reset_index()
            return lov_grouped_df
    
    return pd.DataFrame(columns=['FIELD', 'COUNTRY', 'LOV'])
def cfg_layout(layout_list,project):
    yaml_content = {}
    for file_info in layout_list:
        file_id = file_info.pop('file_id')
        if file_id not in yaml_content:
                yaml_content[file_id]={}
        yaml_content[file_id].update(file_info)

    write_to_yaml_cfg(yaml_content, f"glue/config/cfg_glue_{project}_layout-template.yaml")

def cfg_format(sheet,project,file_list):
    # Catch file_id in file_list
    unique_file_id = set()
    for item in file_list:
        file_id=item["file_id"]
        if file_id not in unique_file_id:
            unique_file_id.add(file_id)

    yaml_content = {}
    default = True
    yaml_content_default = {
        'data_types': {
            'date_format': 'dd/MM/yyyy',
            'decimal_format': '.'
            }
        }
    for row in sheet.iter_rows(min_row=1, values_only=True):
        if not row[1]:
            break  # Stop if File is empty
        if not row[2]:
            continue  # Continue if field is empty

        field = row[2]
        rules = row[4]
        if '*dat*' in field.lower() or '*date*' in field.lower():
            # Define the YAML content
            yaml_content_default = {
                'data_types': {
                    'date_format': rules.replace('"',''),
                    'decimal_format': '.'
                    }
                }
            for file_id in unique_file_id:
                if file_id.lower() in field.lower():
                    # Define the YAML content
                    default = False
                    yaml_content[file_id] = {
                        'data_types': {
                            'date_format': rules.replace('"',''),
                            'decimal_format': '.'
                            }
                        }

    if default: yaml_content=yaml_content_default
    
    write_to_yaml_cfg(yaml_content, f"glue/config/cfg_glue_{project}.yaml")

def cfg_mapping(sheet,project):
    yaml_content = {}
    
    for row in sheet.iter_rows(min_row=5, values_only=True):
        if not row[2]:
            break  # Stop if file_code is empty

        file_code = row[2]
        file_id = row[1]
        yaml_content[file_id] = file_code

    write_to_yaml_cfg(yaml_content, f"glue/config/cfg_glue_{project}_mapping.yaml")


def extract_table_structure(sheet, sheet_name, list_of_value,jv):
    """Extracts the table structure from the given Excel sheet."""
    table_name = sheet_name
    if table_name.endswith("_IN"):
        table_name = table_name[:-3]
    if table_name.endswith("_OUT"):
        table_name = table_name[:-4]
    columns = []
    columns_temp = []
    # Mapping CI types to DataCatalog types
    type_mapping = {
        "alphanumeric": "varchar",
        "numeric": "decimal",
        "date": "date",
        "timestamp": "timestamp"
        # Add other type mappings here
    }
    # Iterate through rows starting from row 9
    for row in sheet.iter_rows(min_row=8, values_only=True):
        if not row[0] or row[0] == "DATE_BATCH_PARTITION":
            break  # Stop if the first cell is empty

        if row[0] == "Field":
            continue
        column_name = row[0]
        column_name = column_name.replace('-','_')
        column_name = column_name.replace('\u00a0', '_')
        column_name = "_".join(column_name.split())
        column_type = row[2] if row[2] else "varchar"  # Default to "Varchar" if column type is empty
        column_precision = str(row[4]).replace('.', ',')  # Convert precision to a string and replace '.' with ','
        column_position = int(row[5])
        column_pk= row[6]
        column_protected = row[3]
        columns_anonymization_rule = row[11]
        column_mandatory='non'
        if row[7] != None and row[7] != '':
            column_mandatory=row[7]
        if column_mandatory == 'oui' or column_mandatory == 'yes': 
            column_mandatory= 'mandatory'

        mapped_type = type_mapping.get(column_type.lower(), column_type.lower())
        mapped_type = mapped_type + f'({column_precision})'
        if 'DATE' in mapped_type.upper():
            mapped_type = 'date'
        if 'TIMESTAMP' in mapped_type.upper():
            mapped_type = 'timestamp'
        
        lov = list_of_value.loc[(list_of_value['FIELD'] == column_name) & (list_of_value['COUNTRY'] == 'ALL'), 'LOV'].tolist()
        if lov == None:
            lov = list_of_value.loc[(list_of_value['FIELD'] == column_name) & (list_of_value['COUNTRY'] == jv.upper()), 'LOV'].tolist()
        columns_temp.append({"position": column_position, "name": column_name, "type": mapped_type, "PrimaryKey":column_pk,"Mandatory":column_mandatory,"Length":column_precision, "LOV": lov, "Protected": column_protected, "AnonymizationRule": columns_anonymization_rule})


    # Sort columns by their position
    columns = sorted(columns_temp, key=lambda x: x["position"])
    
    return {
        "table_name": f"STG_{table_name}",
        "columns": columns,
    }

def get_table_structure_from_excel(workbook, sheet_name, list_of_value,jv):
    """Access the specified sheet and extract the table structure."""
    try:
        sheet = workbook[sheet_name]
        if sheet["A8"].value != "Field" and sheet["A7"].value != "Field":
            return None
        return extract_table_structure(sheet, sheet_name, list_of_value,jv)
    except KeyError:
        print(f"Sheet {sheet_name} not found in the workbook.")
        return None

def write_to_yaml(data, file_path):
    """Write the data to a YAML file."""
    yaml = YAML(typ="safe", pure=True)
    try:
        with open(file_path, "w") as file:
            yaml.dump(data, file)
    except Exception as e:
        print(f"Error writing to {file_path}: {e}")

def write_to_yaml_cfg(data, file_path):
    """Write the data to a YAML file."""
    #yaml = YAML(typ="safe", pure=True)
    try:
        with open(file_path, "w") as file:
            yaml.dump(data, file, allow_unicode=True, default_style="'", default_flow_style=False)
    except Exception as e:
        print(f"Error writing to {file_path}: {e}")

def load_configuration(config_path):
    """Load configuration from the YAML file."""
    try:
        with open(config_path, "r") as file:
            return yaml.safe_load(file)
    except FileNotFoundError:
        print(f"Configuration file {config_path} not found.")
        return None
    except yaml.YAMLError as e:
        print(f"Error loading configuration file: {e}")
        return None

def load_workbook(excel_file_path):
    """Load the Excel workbook."""
    try:
        return openpyxl.load_workbook(excel_file_path)
    except FileNotFoundError:
        print(f"Excel file {excel_file_path} not found.")
        return None
    except openpyxl.utils.exceptions.InvalidFileException as e:
        print(f"Error loading Excel file: {e}")
        return None

def process_files_list(sheet, excel_file_path, jv, periodicity_mapping):
    """Process the 'Files list' sheet to extract file details."""
    files_list_tab = []
    # Dictionnaire pour stocker les periodicities par table_name_output
    periodicity_save_dict = {}
    periodicity_save = []
    table_name_output_save = []
    for row in sheet.iter_rows(min_row=5, values_only=True):
        if not row[2]:
            break  # Stop if file_code is empty

        file_code = row[2]
        table_name_output = row[2]
                
       
                
        if table_name_output.endswith("_IN"):
            table_name_output = table_name_output[:-3]

        periodicity = row[7]
        if periodicity is not None:
            mapped_periodicity = periodicity_mapping.get(periodicity, periodicity)
            
            if table_name_output in periodicity_save_dict:
                if mapped_periodicity not in periodicity_save_dict[table_name_output]:
                    periodicity_save_dict[table_name_output].append(mapped_periodicity)
            else:
                periodicity_save_dict[table_name_output] = [mapped_periodicity]
        else:
            mapped_periodicity = periodicity_mapping.get('d', 'd')            
            if table_name_output in periodicity_save_dict:
                if mapped_periodicity not in periodicity_save_dict[table_name_output]:
                    periodicity_save_dict[table_name_output].append(mapped_periodicity)
            else:
                periodicity_save_dict[table_name_output] = [mapped_periodicity]



        # Exemple d'utilisation : accès aux periodicities pour un table_name_output spécifique
        data_type = row[8]
        file_format = row[10]

        table_name_output_save.append(table_name_output)
        separator = row[11]
        header = row[12]
        footer = row[13]
        quote = row[14]
        update = False
        file_id = row[1]
        layout_code = 'TO_COMPLETED'
        layout_position = 'StartPosition:StopPosition'

        # Check if file_code already exists in the list and update periodicity
        for item in files_list_tab:
            if file_code in item["file_code"]:
                update = True
                break
        if not update :
            files_list_tab.append({
                "file_id": file_id,
                "file_code": file_code,
                "table_name_output": f"STG_{table_name_output.upper()}_OUT",
                "periodicity": periodicity_save_dict[table_name_output],
                "data_type": data_type,
                "format": 'PARQUET',
                "separator": separator,
                "header": header,
                "footer": footer,
                "quote": quote,
                "layout_code": layout_code,
                "layout_position": layout_position
            })
            files_list_tab.append({
                "file_id": file_id,
                "file_code": file_code,
                "table_name_output": f"STG_{table_name_output.upper()}_IN",
                "periodicity": periodicity_save_dict[table_name_output],
                "data_type": data_type,
                "format": file_format,
                "separator": separator,
                "header": header,
                "footer": footer,
                "quote": quote,
                "layout_code": layout_code,
                "layout_position": layout_position
            })
    return files_list_tab

def process_all_sheets(workbook, files_list_tab, jv, yaml_output_path, list_of_value,multi_format,config):
    """Iterate over all sheets in the workbook and process them."""
    layout_list =[]
    for sheet_name in workbook.sheetnames:
        # Next IF must be comment when jv are not in sheetname
        if '_SF' in sheet_name.upper() :
            continue
        if  'TEMPLATE_' in sheet_name.upper() :
            continue
        table_structure = get_table_structure_from_excel(workbook, sheet_name, list_of_value,jv)
        if table_structure:
            file_name = f'STG_{sheet_name.upper()}'
            for item in files_list_tab:
                if file_name in item["table_name_output"] :
                    table_structure = {**table_structure, **item}
                    table_name=table_structure["table_name_output"]
                    yaml_file_path = os.path.join(yaml_output_path, f"{table_name}.yaml")
                    # Make path if not exists
                    if not os.path.exists(yaml_output_path):
                        os.makedirs(yaml_output_path)
                    write_to_yaml(table_structure, yaml_file_path)
                    # layout configuration
                    if multi_format.lower() == 'yes' and table_name.upper().endswith('_IN'):
                        if table_structure["separator"] is not None :
                            layout_separator = table_structure["separator"]
                        else :
                            layout_separator = '|'
                        layout_list.append({'file_id': item["file_id"], re.sub('-IN$',f'-{jv.upper()}-IN',table_name.replace(f'STG_{config["project"].upper()}_','').replace(f'_','-')) :  item["layout_code"], 'layout': item["layout_position"], 'delimiter': layout_separator})

    if multi_format.lower() == 'yes':
        cfg_layout(layout_list,project_path.lower())

if __name__ == "__main__":
    root_folder = os.path.abspath('.')  # Adjust the path as neede
    config_folder = os.path.join(root_folder, 'config')
    config_tech = load_configuration(os.path.join(config_folder, 'config_tech.yaml'))

    if not config_tech:
        exit()
    for file_cfg in config_tech["yaml_cfg_files_path"]:
        config_file = os.path.join(config_folder, file_cfg)
        config = load_configuration(config_file)
        if not config:
            exit()

        excel_file_path = os.path.join(root_folder, config["excel_file_path"]) 
        yaml_output_path = os.path.join(root_folder, config["yaml_path"])
        jv = config["jv"]
        project = config["project"]
        project_path = config["project_path"]
        multi_format = config["multi_layout"]
        
        workbook = load_workbook(excel_file_path)
        if not workbook:
            exit()



        sheet = workbook["Files list"]
        if not sheet:
            print("Sheet 'Files list' not found in the workbook.")
            exit()
        
        periodicity_mapping = {
            "Quotidien": "daily",
            "Hebdomadaire": "weekly",
            "Hebdo": "weekly",
            "Mensuel": "monthly",
            "Annuel": "yearly",
            "w": "weekly",
            "d": "daily",
            "y": "yearly",
            "m": "monthly"
        }

        list_of_value = load_list_of_value(workbook)
        files_list_tab = process_files_list(sheet, excel_file_path, jv, periodicity_mapping)
        process_all_sheets(workbook, files_list_tab, jv, yaml_output_path, list_of_value, multi_format,config)

        # mapping configuration
        cfg_mapping(sheet,project_path.lower())

        # date format configuration
        sheet = workbook["Rules"]
        if not sheet:
            print("Sheet 'Rules' not found in the workbook.")
            exit()
        cfg_format(sheet,project_path.lower(),files_list_tab)
